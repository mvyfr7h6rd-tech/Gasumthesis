"""
Excel export service for GASUM recommendations.
Generates a multi-sheet .xlsx file for a given Recommendation.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from ..models.route import Recommendation


# ── Colour palette ────────────────────────────────────────────────────────────
_HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")   # dark blue header
_HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(name="Calibri", bold=True, size=12)
_BODY_FONT  = Font(name="Calibri", size=10)
_EVEN_FILL  = PatternFill("solid", fgColor="EEF2F7")  # light blue stripe

_RISK_FILLS: Dict[str, PatternFill] = {
    "critical": PatternFill("solid", fgColor="FF4444"),
    "warning":  PatternFill("solid", fgColor="FFA500"),
    "normal":   PatternFill("solid", fgColor="90EE90"),
    "safe":     PatternFill("solid", fgColor="D4EDDA"),
}

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr(ws, row: int, col: int, value: Any) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = _HDR_FILL
    c.font   = _HDR_FONT
    c.border = _BORDER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _cell(ws, row: int, col: int, value: Any, even: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font   = _BODY_FONT
    c.border = _BORDER
    c.alignment = Alignment(horizontal="left", vertical="center")
    if even:
        c.fill = _EVEN_FILL


def _auto_width(ws, min_w: int = 10, max_w: int = 40) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(c.value or "")) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(
            min_w, min(length + 3, max_w)
        )


# ── Sheet 1 – Summary ─────────────────────────────────────────────────────────

def _sheet_summary(wb: Workbook, rec: Recommendation, computation_time: Optional[float]) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"GASUM Biogas Logistics — Plan {rec.id}"
    t.font  = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Generated: {rec.created_at.strftime('%Y-%m-%d %H:%M UTC') if rec.created_at else '—'}"
    ws["A2"].font = Font(name="Calibri", size=10, color="666666")

    rows = [
        ("Plan ID",              rec.id),
        ("Status",               rec.status.value if hasattr(rec.status, "value") else str(rec.status)),
        ("Objective",            rec.objective_function),
        ("Horizon (days)",       rec.horizon_days),
        ("Feasibility level",    rec.feasibility_level),
        ("Solution risk score",  f"{rec.solution_risk_score:.1f}/10" if rec.solution_risk_score is not None else "—"),
        ("Computation time (s)", f"{computation_time:.2f}" if computation_time else "—"),
        ("", ""),
        ("Total distance (km)",  round(rec.total_distance_km, 1)),
        ("Transport cost (EUR)", round(rec.transport_cost_eur, 2)),
        ("Handling cost (EUR)",  round(rec.handling_cost_eur, 2)),
        ("Contingency cost (EUR)", round(rec.total_cost_eur - rec.transport_cost_eur - rec.handling_cost_eur, 2)),
        ("TOTAL COST (EUR)",     round(rec.total_cost_eur, 2)),
        ("", ""),
        ("Sites served",         rec.sites_served),
        ("Critical sites addressed", rec.critical_sites_addressed),
        ("Total MWh moved",      round(rec.total_mwh_moved, 2) if rec.total_mwh_moved else 0),
        ("EUR / MWh",            round(rec.eur_per_mwh, 2) if rec.eur_per_mwh else "—"),
        ("", ""),
        ("Explanation",          rec.explanation or "—"),
    ]

    for i, (label, value) in enumerate(rows, start=4):
        if not label:
            continue
        ws.cell(row=i, column=1, value=label).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=i, column=1).border = Border(bottom=_THIN)
        c = ws.cell(row=i, column=2, value=value)
        c.font   = _BODY_FONT
        c.border = Border(bottom=_THIN)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 35


# ── Sheet 2 – Routes ──────────────────────────────────────────────────────────

def _sheet_routes(wb: Workbook, rec: Recommendation) -> None:
    ws = wb.create_sheet("Routes")
    ws.sheet_view.showGridLines = False

    headers = [
        "Day", "Truck", "Seq", "Site", "Operation",
        "Containers Dropped", "Containers Picked",
        "Cumul. Distance (km)", "Load Full After", "Load Empty After",
        "Arrival Time (h)", "Service Time (h)",
    ]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 28

    row = 2
    for route in rec.routes:
        for stop in route.stops:
            op = stop.swap_operation
            dropped_ids = ", ".join(op.containers_dropped) if op and op.containers_dropped else ""
            picked_ids  = ", ".join(op.containers_picked)  if op and op.containers_picked  else ""

            has_activity = op and (op.containers_dropped or op.containers_picked)
            if has_activity:
                op_text = f"drop={len(op.containers_dropped or [])} pick={len(op.containers_picked or [])}"
            else:
                op_text = "transit"

            even = (row % 2 == 0)
            vals = [
                route.day_index,
                route.truck_id,
                stop.sequence,
                stop.site_name or stop.site_id,
                op_text,
                dropped_ids or "—",
                picked_ids  or "—",
                round(stop.cumulative_distance_km, 1),
                stop.load_full_after,
                stop.load_empty_after,
                round(stop.arrival_time_hours, 2),
                round(stop.service_time_hours, 2),
            ]
            for col, val in enumerate(vals, 1):
                _cell(ws, row, col, val, even)
            row += 1

    _auto_width(ws)


# ── Sheet 3 – Container Moves ─────────────────────────────────────────────────

def _sheet_container_moves(wb: Workbook, rec: Recommendation) -> None:
    ws = wb.create_sheet("Container Moves")
    ws.sheet_view.showGridLines = False

    headers = ["Day", "Truck", "Bay", "Container ID", "From Site", "To Site", "Move Type"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 28

    row = 2
    # Use container_moves field if present, otherwise derive from routes
    moves = getattr(rec, "container_moves", None) or []
    if moves:
        for move in moves:
            even = (row % 2 == 0)
            vals = [
                getattr(move, "day_index", "—"),
                getattr(move, "truck_id", "—"),
                getattr(move, "bay_id", "—"),
                getattr(move, "container_serial", "—"),
                getattr(move, "from_site_name", getattr(move, "from_site_id", "—")),
                getattr(move, "to_site_name",   getattr(move, "to_site_id",   "—")),
                getattr(move, "move_type", "—"),
            ]
            for col, val in enumerate(vals, 1):
                _cell(ws, row, col, val, even)
            row += 1
    else:
        # Derive from route stops swap operations
        for route in rec.routes:
            stops = route.stops
            for i, stop in enumerate(stops):
                op = stop.swap_operation
                if not op:
                    continue
                for cid in (op.containers_dropped or []):
                    even = (row % 2 == 0)
                    # "dropped" = delivered to this site; came from previous stop
                    from_site = stops[i - 1].site_name if i > 0 else "?"
                    vals = [route.day_index, route.truck_id, "—", cid, from_site,
                            stop.site_name or stop.site_id, "full delivery"]
                    for col, val in enumerate(vals, 1):
                        _cell(ws, row, col, val, even)
                    row += 1
                for cid in (op.containers_picked or []):
                    even = (row % 2 == 0)
                    vals = [route.day_index, route.truck_id, "—", cid,
                            stop.site_name or stop.site_id, "→ next stop", "empty return"]
                    for col, val in enumerate(vals, 1):
                        _cell(ws, row, col, val, even)
                    row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="No container moves recorded").font = _BODY_FONT

    _auto_width(ws)


# ── Sheet 4 – Warnings & Feedback ─────────────────────────────────────────────

def _sheet_warnings(wb: Workbook, rec: Recommendation) -> None:
    ws = wb.create_sheet("Warnings & Feedback")
    ws.sheet_view.showGridLines = False

    _hdr(ws, 1, 1, "Type")
    _hdr(ws, 1, 2, "Message")
    ws.row_dimensions[1].height = 28

    row = 2
    for w in (rec.warnings or []):
        even = (row % 2 == 0)
        _cell(ws, row, 1, "WARNING", even)
        _cell(ws, row, 2, str(w), even)
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10, color="C00000")
        row += 1

    for fb in (rec.solution_feedback or []):
        level = str(fb.get("level", "info")).upper() if isinstance(fb, dict) else "INFO"
        msg   = fb.get("message", str(fb)) if isinstance(fb, dict) else str(fb)
        even = (row % 2 == 0)
        _cell(ws, row, 1, level, even)
        _cell(ws, row, 2, msg, even)
        row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="No warnings").font = _BODY_FONT

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 80


# ── Public API ────────────────────────────────────────────────────────────────

def build_recommendation_excel(
    rec: Recommendation,
    computation_time: Optional[float] = None,
) -> bytes:
    """
    Build a multi-sheet Excel workbook for *rec* and return raw bytes.
    Sheets: Summary | Routes | Container Moves | Warnings & Feedback
    """
    wb = Workbook()
    _sheet_summary(wb, rec, computation_time)
    _sheet_routes(wb, rec)
    _sheet_container_moves(wb, rec)
    _sheet_warnings(wb, rec)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
